import openpyxl
import unicodedata
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pyexcel_ods3 import get_data  # to open ods files
from dataclasses import dataclass
from json import encoder


# Nodo para guardar los datos de cada entrada de los Excel de campus virtual y UXXI
@dataclass
class Registro:
    nombre: str
    nombre_anomalia_uxxi: str  # Añadido por un problema en el excel uxxi con los estudiantes de un solo apellido, añade un espacio antes de la coma de separación del nombre (ej. Mateusz , XXXXXX)
    nota_cv: float | None
    nota_uxxi: float | None
    nota_exp: float | None


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestión de Actas - Campus Virtual a UXXI v1.00")
        self.root.geometry("1100x600")

        # Lista de objetos para almacenar la información
        # Cargaremos los datos de los Excel de CV y UXXI en esta lista
        # Distinguiremos por los valores None en nota_cv y nota_uxxi de que fichero Excel vienen los datos
        # None en nota_cv significa que no está en el excel de CV
        # None en nota_uxxi significa que no está en el excel de UXXI
        self.lista_registros = []

        self.Fichero_Export = ""
        self._syncing = False

        # Frame principal
        main_frame = tk.Frame(root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # Configurar grid weight para redimensionamiento responsivo
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(1, weight=1)

        # =========================================================
        # SECCIÓN CAMPUS VIRTUAL (Izquierda)
        # =========================================================
        lbl_cv = tk.Label(main_frame, text="CAMPUS VIRTUAL", font=("Arial", 12, "bold"))
        lbl_cv.grid(row=0, column=0, pady=(0, 10))

        # Listado (Treeview) Izquierda
        columns_cv = ("NOMBRE", "NOTA_CV")
        self.tree_cv = ttk.Treeview(main_frame, columns=columns_cv, show="headings")
        self.tree_cv.heading("NOMBRE", text="NOMBRE")
        self.tree_cv.heading("NOTA_CV", text="NOTA_CV")

        self.tree_cv.column("NOMBRE", width=350, stretch=True)
        self.tree_cv.column("NOTA_CV", width=100, stretch=False, anchor=tk.CENTER)

        # Grid para el Treeview izquierdo
        self.tree_cv.grid(row=1, column=0, sticky="nsew", padx=(0, 10))

        # Scrollbar para el Treeview izquierdo
        self.scrollbar_cv = ttk.Scrollbar(
            main_frame, orient=tk.VERTICAL, command=self.sync_yview
        )
        self.tree_cv.configure(yscrollcommand=self.sync_yscroll_set)
        self.scrollbar_cv.grid(row=1, column=0, sticky="nse", padx=(0, 10))

        # Botón Cargar Datos CV
        btn_cv = tk.Button(
            main_frame,
            text="(1) CARGAR DATOS CV",
            command=self.cargar_datos_cv,
            height=2,
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10, "bold"),
        )
        btn_cv.grid(row=2, column=0, pady=(15, 5), padx=(0, 10), sticky="ew")

        # Botón Proceso
        btn_proceso = tk.Button(
            main_frame,
            text="(3) PROCESAR",
            command=self.procesar_datos,
            height=2,
            bg="#FF9800",
            fg="white",
            font=("Arial", 10, "bold"),
        )
        btn_proceso.grid(row=3, column=0, pady=(5, 0), padx=(0, 10), sticky="ew")

        # =========================================================
        # SECCIÓN UXXI (Derecha)
        # =========================================================
        lbl_uxxi = tk.Label(main_frame, text="UXXI", font=("Arial", 12, "bold"))
        lbl_uxxi.grid(row=0, column=1, pady=(0, 10))

        # Listado (Treeview) Derecha
        columns_uxxi = ("NOMBRE", "NOTA_UXXI", "NOTA_EXP")
        self.tree_uxxi = ttk.Treeview(main_frame, columns=columns_uxxi, show="headings")
        self.tree_uxxi.heading("NOMBRE", text="NOMBRE")
        self.tree_uxxi.heading("NOTA_UXXI", text="NOTA_UXXI")
        self.tree_uxxi.heading("NOTA_EXP", text="NOTA_EXP")

        self.tree_uxxi.column("NOMBRE", width=300, stretch=True)
        self.tree_uxxi.column("NOTA_UXXI", width=120, stretch=False, anchor=tk.CENTER)
        self.tree_uxxi.column("NOTA_EXP", width=120, stretch=False, anchor=tk.CENTER)

        # Grid para el Treeview derecho
        self.tree_uxxi.grid(row=1, column=1, sticky="nsew", padx=(10, 0))

        # Scrollbar para el Treeview derecho
        self.scrollbar_uxxi = ttk.Scrollbar(
            main_frame, orient=tk.VERTICAL, command=self.sync_yview
        )
        self.tree_uxxi.configure(yscrollcommand=self.sync_yscroll_set)
        self.scrollbar_uxxi.grid(row=1, column=1, sticky="nse", padx=(10, 0))

        # Botón Cargar Datos UXXI
        btn_uxxi = tk.Button(
            main_frame,
            text="(2) CARGAR DATOS UXXI",
            command=self.cargar_datos_uxxi,
            height=2,
            bg="#2196F3",
            fg="white",
            font=("Arial", 10, "bold"),
        )
        btn_uxxi.grid(row=2, column=1, pady=(15, 5), padx=(10, 0), sticky="ew")

        # Botón Exportar
        btn_exportar = tk.Button(
            main_frame,
            text="(4) EXPORTAR",
            command=self.exportar_datos,
            height=2,
            bg="#9C27B0",
            fg="white",
            font=("Arial", 10, "bold"),
        )
        btn_exportar.grid(row=3, column=1, pady=(5, 0), padx=(10, 0), sticky="ew")

        # Configuración de colores para las filas
        self.tree_cv.tag_configure("naranja", background="#FFB74D")
        self.tree_uxxi.tag_configure("naranja", background="#FFB74D")

        self.tree_cv.tag_configure("rojo", background="#E57373")
        self.tree_uxxi.tag_configure("rojo", background="#E57373")

        self.tree_cv.tag_configure("verde", background="#81C784")
        self.tree_uxxi.tag_configure("verde", background="#81C784")

        self.tree_cv.tag_configure("marron", background="#8D6E63", foreground="white")
        self.tree_uxxi.tag_configure("marron", background="#8D6E63", foreground="white")

    def sync_yview(self, *args):
        self.tree_cv.yview(*args)
        self.tree_uxxi.yview(*args)

    def sync_yscroll_set(self, *args):
        self.scrollbar_cv.set(*args)
        self.scrollbar_uxxi.set(*args)
        if getattr(self, "_syncing", False):
            return
        self._syncing = True
        try:
            self.tree_cv.yview_moveto(args[0])
            self.tree_uxxi.yview_moveto(args[0])
        finally:
            self._syncing = False

    # =========================================================
    # SECCIÓN PROCESAR DATOS
    # =========================================================
    # Procesar los datos de los listados de Campus Virtual y UXXI, y fusionarlos
    def procesar_datos(self):
        # Vaciamos la lista de registros para ir añadiendo los estudiantes según recorremos los listados
        self.lista_registros.clear()
        # Creamos un diccionario para gestionar los duplicados en ambos listados
        registros_dict = {}

        # Función auxiliar para convertir las notas a float
        def parse_float(val):
            if not val:
                return 0.0
            try:
                if isinstance(val, (int, float)):
                    return float(val)
                # Reemplazar coma por punto por si acaso viene con formato europeo
                val_str = str(val).replace(",", ".").strip()
                return float(val_str)
            except ValueError:
                return 0.0

        # 1. Leer datos del listado de la izquierda (del Excel de Campus Virtual) y guardarlos en un diccionario
        for child in self.tree_cv.get_children():
            # Sacamos los datos (nombre y nota del campus) de cada línea del listado de Campus Virtual
            nombre, nota_cv = self.tree_cv.item(child, "values")

            # Comprobar si el nombre ya existe en el diccionario (=> que estaba duplicado en el Excel del Campus Virtual)
            if nombre in registros_dict:
                raise Exception(
                    f"Hay dos entradas de este estudiante en el Excel de Campus Virtual: {nombre}"
                )

            # Guardamos los datos en el diccionario para fusionar los datos de Campus Virtual con los de UXXI
            registros_dict[nombre] = Registro(
                nombre=nombre,
                nombre_anomalia_uxxi=None,
                nota_cv=parse_float(nota_cv),
                nota_uxxi=None,
                nota_exp=None,
            )

        # 2. Leer datos del listado de la derecha (del Excel de UXXI) y añadir/fusionar con los anteriores
        for child in self.tree_uxxi.get_children():
            # Sacamos los datos de una línea del listado de UXXI
            nombre, nota_uxxi, nota_exp = self.tree_uxxi.item(child, "values")
            val_uxxi = parse_float(nota_uxxi)
            val_exp = parse_float(nota_exp)

            # En este caso no vamos a comprobar que no hay duplicados en el Excel de UXXI, como hicimos con el de Campus Virtual

            # Comprobamos si el nombre ya existe en el diccionario (=> ya lo cargamos que estaba duplicado en el Excel del Campus Virtual)
            if nombre in registros_dict:
                registros_dict[nombre].nota_uxxi = val_uxxi
                registros_dict[nombre].nota_exp = val_exp
            else:  # El nombre y apellidos, tal como vienen desde el fichero UXXI, no están en el fichero de campus virtual
                if (
                    " , " in nombre
                ):  # Comprobamos que sea candidato a anomalia de UXXI viendo si contiene " , "
                    # Comprobamos si este es un caso de la anomalía de UXXI con estudiantes de un sólo apellido que mete un espacio extra antes de la coma
                    nombre_formato_cv = nombre.replace(
                        " , ", ", "
                    )  # quitamos el espacio de antes de la coma, y si así está entre los que cargamos de campus virtual lo dejamos señalizado en nombre_anomalia_uxxi para poder exportar la nota correctamente
                    if nombre_formato_cv in registros_dict:
                        registros_dict[nombre_formato_cv].nota_uxxi = val_uxxi
                        registros_dict[nombre_formato_cv].nota_exp = val_exp
                        registros_dict[nombre_formato_cv].nombre_anomalia_uxxi = nombre
                    else:  # en este caso ya sólo queda que es un estudiante que sólo está en UXXI y no hay datos de campus virtual (posible caso de estudiante que ha perdido la evaluación continua, y este sistema le asignará una nota de 0.0)
                        registros_dict[nombre] = Registro(
                            nombre=nombre,
                            nombre_anomalia_uxxi=None,
                            nota_cv=None,
                            nota_uxxi=val_uxxi,
                            nota_exp=val_exp,
                        )
                else:
                    registros_dict[nombre] = Registro(
                        nombre=nombre,
                        nombre_anomalia_uxxi=None,
                        nota_cv=None,
                        nota_uxxi=val_uxxi,
                        nota_exp=val_exp,
                    )

        # 3. Pasar diccionario a la lista y ordenar por nombre alfabéticamente ignorando tildes y mayúsculas
        self.lista_registros = list(registros_dict.values())

        # Función para ordenar la lista de forma insensible a mayúsculas/minúsculas y acentos
        def normalizar_para_ordenar(texto):
            # Elimina acentos/diacríticos y pasa a minúsculas
            if not texto:
                return ""
            texto_nfkd = unicodedata.normalize("NFKD", texto)
            return "".join(
                [c for c in texto_nfkd if not unicodedata.combining(c)]
            ).lower()

        # Ordenamos la lista
        self.lista_registros.sort(key=lambda req: normalizar_para_ordenar(req.nombre))

        messagebox.showinfo(
            "Proceso Completo",
            f"Se han cruzado y ordenado {len(self.lista_registros)} registros correctamente.",
        )

        # Vaciamos los treeview (listados de la izquierda y derecha)
        for item in self.tree_cv.get_children():
            self.tree_cv.delete(item)
        for item in self.tree_uxxi.get_children():
            self.tree_uxxi.delete(item)

        # Recargamos los listados para que queden matcheados e identificados con diferentes colores las diferentes casuisticas entre los listados izquierdo y derecho
        # 4. Volcar la lista ordenada a los Treeview
        for reg in self.lista_registros:
            if reg.nota_cv is None:
                # Falta en Campus Virtual -> Naranja y nota_exp a 0.0 (candidatos a haber perdido la evaluación continua)
                reg.nota_exp = 0.0
                nota_uxxi_str = (
                    f"{reg.nota_uxxi:.1f}" if reg.nota_uxxi is not None else "-"
                )
                nota_exp_str = f"{reg.nota_exp:.1f}"

                val_cv = ("", "")
                val_uxxi = (reg.nombre, nota_uxxi_str, nota_exp_str)
                tag = "naranja"

            elif reg.nota_uxxi is None:
                # Falta en UXXI -> Rojo y nota_exp a None (estudiante que aparece en Campus Virtual pero no en UXXI, probable caso de error en algún caracter del nombre/apellidos, seguramente algúna diferencia en acentos entre el nombre en campus virtual y en UXXI)
                reg.nota_exp = None
                nota_cv_str = f"{reg.nota_cv:.1f}" if reg.nota_cv is not None else "-"
                nota_exp_str = "-"

                val_cv = (reg.nombre, nota_cv_str)
                val_uxxi = ("", "", "")
                tag = "rojo"

            elif (
                reg.nota_uxxi is not None
                and reg.nota_uxxi != 0
                and reg.nota_cv != reg.nota_uxxi
            ):
                # El estudiante está en Campus Virtual y en UXXI, y tiene registrada una nota en UXXI que no es 0 ni None (suponemos que asignada previamente de forma explícita), pero no coincide con la de campus virtual
                # En este caso propondremos para exportar la nota de Campus Virtual, pero marcaremos el registro en marrón para revisar antes de exportar
                reg.nota_exp = reg.nota_cv
                nota_cv_str = f"{reg.nota_cv:.1f}"
                nota_uxxi_str = f"{reg.nota_uxxi:.1f}"
                nota_exp_str = f"{reg.nota_exp:.1f}"

                val_cv = (reg.nombre, nota_cv_str)
                val_uxxi = (reg.nombre, nota_uxxi_str, nota_exp_str)
                tag = "marron"

            else:
                # Están en ambos -> Verde
                reg.nota_exp = reg.nota_cv
                nota_cv_str = f"{reg.nota_cv:.1f}"
                nota_uxxi_str = f"{reg.nota_uxxi:.1f}"
                nota_exp_str = f"{reg.nota_cv:.1f}"

                val_cv = (reg.nombre, nota_cv_str)
                val_uxxi = (reg.nombre, nota_uxxi_str, nota_exp_str)
                tag = "verde"

            # Insertar en el Treeview de la izquierda (Campus Virtual)
            self.tree_cv.insert("", tk.END, values=val_cv, tags=(tag,))

            # Insertar en el Treeview de la derecha (UXXI)
            self.tree_uxxi.insert("", tk.END, values=val_uxxi, tags=(tag,))

        # Ubicar los listados al principio (scroll arriba)
        self.tree_cv.yview_moveto(0)
        self.tree_uxxi.yview_moveto(0)

    def exportar_datos(self):
        if self.Fichero_Export != "":
            messagebox.showinfo(
                "Exportando Calificaciones a UXXI",
                f"Fichero: {self.Fichero_Export}\n",
            )
        else:
            messagebox.showinfo(
                "Exportando Calificaciones a UXXI",
                "No se ha seleccionado ningún fichero Excel de UXXI",
            )
            return

        try:
            wb = openpyxl.load_workbook(self.Fichero_Export)
            sheet = wb.active

            for reg in self.lista_registros:
                if reg.nota_exp is None:
                    continue

                if reg.nombre_anomalia_uxxi is not None:
                    nombre_buscar = str(reg.nombre_anomalia_uxxi).strip()
                else:
                    nombre_buscar = str(reg.nombre).strip()
                fila_encontrada = None

                for row in sheet.iter_rows(min_row=2):
                    valor_nombre = row[0].value
                    if valor_nombre is None:
                        continue
                    if str(valor_nombre).strip() == nombre_buscar:
                        fila_encontrada = row
                        break

                if fila_encontrada is None:
                    raise Exception(
                        f"No se localiza el nombre del estudiante '{nombre_buscar}' en el fichero UXXI"
                    )
                # En el Excel UXXI las notas utilizan coma decimal y una sola cifra decimal
                # valor_nota = f"{reg.nota_exp:.1f}".replace(".", ",")
                fila_encontrada[2].value = reg.nota_exp

            wb.save(self.Fichero_Export)
            messagebox.showinfo(
                "Exportación completada",
                "Se ha terminado la exportación.",
            )
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"Error al exportar las calificaciones:\n{str(e)}",
            )

    def obtener_indice_columna(self, encabezados, nombre_columna):
        for i, encabezado in enumerate(encabezados):
            if encabezado is not None and str(encabezado).strip() == nombre_columna:
                return i
        return -1

    def obtener_indice_columna_ods(self, encabezados, nombre_columna):
        for i, encabezado in enumerate(encabezados):
            if encabezado is not None and str(encabezado).strip() == nombre_columna:
                return i
        return -1

    def obtener_indice_columna_xlsx(self, encabezados, nombre_columna):
        for i, encabezado in enumerate(encabezados):
            if encabezado is not None and str(encabezado).strip() == nombre_columna:
                return i
        return -1

    def cargar_datoscv_ods(self, filepath):
        try:
            # Cargar archivo ODS
            data = get_data(filepath)
            # Obtener la primera hoja como lista de listas
            sheet = next(iter(data.values()))
            # Si no hay datos en el fichero => Excepción
            if not sheet or len(sheet) == 0:
                raise Exception("El archivo ODS está vacío.")

            # Obtenemos la posición de la columna con la nota final de la asignatura
            # Primero intentamos con el nombre de la columna en español "Total del curso (Real)"
            idx_nota = self.obtener_indice_columna(sheet[0], "Total del curso (Real)")
            # Si no se encuentra, intentamos con el nombre en inglés "Course total (Real)"
            if idx_nota == -1:
                idx_nota = self.obtener_indice_columna(sheet[0], "Course total (Real)")
                if idx_nota == -1:  # Si no se encuentra con ningún nombre => Excepción
                    raise Exception("No se ha encontrado la columna con la nota final de la asignatura.")

            # Limpiar listado actual
            for item in self.tree_cv.get_children():
                self.tree_cv.delete(item)

            # Iterar desde la fila 2 (índice 1) para omitir encabezados
            for row in sheet[1:]:
                # Evitar procesar filas completamente vacías
                if not any(row):
                    continue

                # Columna A -> índice 0 | Columna B -> índice 1
                # Obtenemos primero el nombre y después los apellidos del estudiante
                col_A = (
                    str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
                )
                col_B = (
                    str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                )

                # Concatenar nombre y apellidos para que coincida con el formato de UXXI
                nombre = f"{col_B}, {col_A}"
                if not col_B and not col_A:
                    nombre = ""

                # Extraer nota final de la asignatura de la columna que hemos localizado al principio
                nota_cv = (
                    str(row[idx_nota]).strip()
                    if len(row) > idx_nota and row[idx_nota] is not None
                    else ""
                )

                # Insertar solo si hay datos relevantes
                if nombre or nota_cv:
                    self.tree_cv.insert("", tk.END, values=(nombre, nota_cv))

        except Exception as e:
            messagebox.showerror(
                "Error", f"Error al cargar el fichero ODS de Campus Virtual:\n{str(e)}"
            )

    def cargar_datoscv_xlsx(self, filepath):
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = wb.active

            try:
                encabezados = next(
                    sheet.iter_rows(min_row=1, max_row=1, values_only=True)
                )
            except StopIteration:
                raise Exception("El archivo Excel está vacío.")

            # Obtenemos la posición de la columna con la nota final de la asignatura
            # Primero intentamos con el nombre de la columna en español "Total del curso (Real)"
            idx_nota = self.obtener_indice_columna_xlsx(encabezados, "Total del curso (Real)")
            # Si no se encuentra, intentamos con el nombre en inglés "Course total (Real)"
            if idx_nota == -1:
                idx_nota = self.obtener_indice_columna_xlsx(encabezados, "Course total (Real)")
                if idx_nota == -1:  # Si no se encuentra con ningún nombre => Excepción
                    raise Exception("No se ha encontrado la columna con la nota final de la asignatura.")

            # Limpiar listado actual
            for item in self.tree_cv.get_children():
                self.tree_cv.delete(item)

            # Iterar empezando desde la fila 2 para omitir la primera fila de títulos
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Evitar procesar filas completamente vacías
                if not any(row):
                    continue

                # Columna A corresponde al índice 0
                # Columna B corresponde al índice 1
                col_A = (
                    str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
                )
                col_B = (
                    str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                )

                # Concatenar la columna B con los caracteres ", " y con la columna A
                nombre = f"{col_B}, {col_A}"
                # Si ambas están vacías, dejamos el nombre en blanco para limpiar el formato
                if not col_B and not col_A:
                    nombre = ""

                # Extraer NOTA_CV de la columna que hemos localizado al principio
                nota_cv = (
                    str(row[idx_nota]).strip()
                    if len(row) > idx_nota and row[idx_nota] is not None
                    else ""
                )

                # Solo añadir al listado si pudimos extraer al menos el nombre o la nota
                if nombre or nota_cv:
                    self.tree_cv.insert("", tk.END, values=(nombre, nota_cv))

        except Exception as e:
            messagebox.showerror(
                "Error", f"Error al cargar el Excel de Campus Virtual:\n{str(e)}"
            )

    # Cargar datos del fichero de Campus Virtual en el listado de la izquierda
    def cargar_datos_cv(self):
        # Pedimos al usuario que seleccione el fichero de Campus Virtual
        filepath = filedialog.askopenfilename(
            title="Seleccionar fichero Campus Virtual (.ods o .xlsx/.xls)",
            filetypes=[("Archivos ODS y Excel", "*.ods *.xlsx *.xls")],
        )
        if not filepath:
            return

        if filepath.lower().endswith(".ods"):
            self.cargar_datoscv_ods(filepath)
        elif filepath.lower().endswith((".xlsx", ".xls")):
            self.cargar_datoscv_xlsx(filepath)
        else:
            messagebox.showerror("Error", "Formato de archivo no soportado.")

    # Cargar datos del Excel de UXXI en el listado de la derecha
    def cargar_datos_uxxi(self):
        # Pedimos al usuario que seleccione el fichero Excel de UXXI
        filepath = filedialog.askopenfilename(
            title="Seleccionar fichero Excel UXXI",
            filetypes=[("Archivos Excel", "*.xlsx *.xls")],
        )
        # Si no se selecciona ningún fichero, salimos de la función
        if not filepath:
            return

        try:
            # Guardamos la ruta del fichero en la variable self.Fichero_Export para usarla en la exportación
            self.Fichero_Export = filepath
            # Cargamos el fichero Excel seleccionado por el usuario
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = wb.active

            # Limpiar listado actual
            for item in self.tree_uxxi.get_children():
                self.tree_uxxi.delete(item)

            # Iterar empezando desde la fila 2 para omitir la primera fila de títulos
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Evitar procesar filas completamente vacías
                if not any(row):
                    continue

                # Extraer la columna A (índice 0)
                # contiene el nombre y apellidos del estudiante separados por coma
                nombre = (
                    str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
                )

                # NOTA_UXXI contiene la nota final de la asignatura si ya se había subido a UXXI
                # si está vacío significa que no se ha subido la nota a UXXI
                nota_uxxi = (
                    str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                )
                # NOTA_EXP contendrá la nota final de la asignatura que se exportará a UXXI cuando hayamos procesado los datos con rl botón "(3) Procesar"
                nota_exp = ""

                # Solo añadir al listado si extrajimos un nombre (podría ser una fila en blanco al final)
                if nombre:
                    self.tree_uxxi.insert(
                        "", tk.END, values=(nombre, nota_uxxi, nota_exp)
                    )

        except Exception as e:
            messagebox.showerror(
                "Error", f"Error al cargar el Excel de UXXI:\n{str(e)}"
            )


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)

    # Configuración de estilo global (tema de Tcl/Tk)
    style = ttk.Style()
    style.theme_use("clam")
    style.configure("Treeview.Heading", font=("Arial", 10, "bold"))
    style.configure("Treeview", font=("Arial", 10), rowheight=25)

    root.mainloop()
